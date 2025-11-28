# insert_pp_species_match.py
import pandas as pd
import unicodedata
import os
import re
from openpyxl import load_workbook

# CONFIG
file_bio = "docs/dados_biologicos.xlsx"
sheet_conn = "Conexão com Políticas Públicas"
sheet_species = "Informações sobre as espécies "
file_pp = "docs/public_policies.xlsx"
out_dir = "sql"
os.makedirs(out_dir, exist_ok=True)

# util: normalizar texto
def normalize(s):
    if pd.isna(s) or s is None:
        return None
    s = str(s).strip()
    if s == "":
        return None
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

# detecta se string é intervalo Excel tipo "E6:E16" ou "E6"
_range_re = re.compile(r'^[A-Za-z]+[0-9]+(:[A-Za-z]+[0-9]+)?$')

# converte letra da coluna Excel para índice 0-based
def col_letter_to_index(letters):
    letters = letters.upper()
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n - 1

# lê valores de um intervalo na sheet de espécies usando openpyxl (preserva exatamente o que está na planilha)
def read_excel_range_values(path, sheet_name, cell_range):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    cells = ws[cell_range]
    values = []
    # cells pode ser 1xN ou NxM; iterar e coletar valores em ordem
    if hasattr(cells, "__iter__"):
        for row in cells:
            if hasattr(row, "__iter__"):
                for c in row:
                    values.append(c.value)
            else:
                values.append(row.value)
    else:
        values.append(cells.value)
    return values

# CARREGA dados
print("Carregando planilhas...")
df_conn = pd.read_excel(file_bio, sheet_name=sheet_conn)
df_pp = pd.read_excel(file_pp)
# tentativa de ler sheet de espécies com fallback de auto cabeçalho
try:
    df_sp = pd.read_excel(file_bio, sheet_name=sheet_species)
except Exception:
    df_sp = pd.read_excel(file_bio, sheet_name=sheet_species, header=0)

# construir map title -> resourceID a partir do arquivo public_policies.xlsx
map_pp = {}
for _, r in df_pp.iterrows():
    t = r.get("title") or r.get("Title") or r.get("titulo")
    rid = r.get("resourceID") or r.get("resourceId") or r.get("resourceid")
    if pd.isna(t) or pd.isna(rid):
        continue
    key = normalize(t)
    try:
        map_pp[key] = int(rid)
    except:
        continue

# construir maps de espécies: nome científico, vernacular e também speciesID direto
map_species_by_name = {}   # normalized name -> speciesID
map_species_by_id = set()  # set of existing speciesID
# normalizar nomes das colunas do df_sp para localizar colunas úteis
cols_norm = {normalize(c): c for c in df_sp.columns}

# descobrir qual coluna tem speciesID na aba de espécies
species_id_col = None
for key in ["speciesid", "species_id", "id", "idSpecies"]:
    if key in cols_norm:
        species_id_col = cols_norm[key]
        break
# nome científico e vernacular
candidate_name_cols = []
for cand in ["scientificname","scientific name","nome cientifico","scientific_name","name"]:
    if cand in cols_norm:
        candidate_name_cols.append(cols_norm[cand])
for cand in ["species","espécie","vernacularname","nome popular","common name","vernacular name"]:
    if cand in cols_norm and cols_norm[cand] not in candidate_name_cols:
        candidate_name_cols.append(cols_norm[cand])

# build maps
for _, r in df_sp.iterrows():
    sid = None
    if species_id_col:
        sid = r.get(species_id_col)
        if pd.isna(sid):
            sid = None
        else:
            try:
                sid = int(sid)
                map_species_by_id.add(sid)
            except:
                sid = None
    # map by names
    for ncol in candidate_name_cols:
        val = r.get(ncol)
        if pd.notna(val):
            k = normalize(val)
            if k:
                # if sid present map to sid; else store placeholder None and resolve later by name
                if sid:
                    map_species_by_name[k] = sid
                else:
                    # store sentinel -1 for now if id not present; later we won't be able to insert without speciesID
                    map_species_by_name[k] = map_species_by_name.get(k, None)

# Prepare outputs
inserts = []
erros = []

# função auxiliar para obter speciesIDs a partir de:
# - coluna name (equality with biologicalLink) OR
# - intervalo excel com nomes/ids OR
# - lista de nomes/ids no próprio cell_value splitted by //,;,,
def resolve_species_ids(cell_spec_col, biologicalLink):
    """
    cell_spec_col: string from speciesInformationColumn cell (can be column name OR Excel range OR list)
    biologicalLink: value to match inside the specified column (if speciesInformationColumn is a column name)
    returns list of speciesID ints (may be empty)
    """
    result_ids = set()

    if cell_spec_col is None:
        return []

    cell_spec_col = str(cell_spec_col).strip()
    # Case 1: it's a column name in species sheet
    nk = normalize(cell_spec_col)
    if nk in cols_norm:
        real_col = cols_norm[nk]
        # match rows where normalize(df_sp[real_col]) == normalize(biologicalLink)
        target_val = normalize(biologicalLink)
        if target_val is None:
            return []
        for _, rr in df_sp.iterrows():
            v = rr.get(real_col)
            if pd.isna(v):
                continue
            if normalize(v) == target_val:
                # try get speciesID
                if species_id_col:
                    sid = rr.get(species_id_col)
                    if pd.notna(sid):
                        try:
                            result_ids.add(int(sid))
                        except:
                            pass
                else:
                    # try to match by name columns if no speciesID present
                    for ncol in candidate_name_cols:
                        nv = rr.get(ncol)
                        if pd.notna(nv) and normalize(nv) == target_val:
                            # cannot get id - skip
                            pass
        return sorted(result_ids)

    # Case 2: looks like an excel range (E6:E16) or single cell like E6
    if _range_re.match(cell_spec_col):
        try:
            values = read_excel_range_values(file_bio, sheet_species, cell_spec_col)
            # values likely are names or ids; try to interpret each
            for val in values:
                if val is None:
                    continue
                sval = str(val).strip()
                if sval.isdigit():
                    try:
                        result_ids.add(int(sval))
                        continue
                    except:
                        pass
                # try name lookup
                k = normalize(sval)
                sid = map_species_by_name.get(k)
                if sid:
                    result_ids.add(int(sid))
            return sorted(result_ids)
        except Exception as e:
            return []

    # Case 3: the cell contains a list of names/ids separated by // or ; or ,
    parts = [p.strip() for p in re.split(r'//|;|,|\n', cell_spec_col) if p.strip() != ""]
    for p in parts:
        if p.isdigit():
            try:
                result_ids.add(int(p))
                continue
            except:
                pass
        k = normalize(p)
        sid = map_species_by_name.get(k)
        if sid:
            result_ids.add(int(sid))
    return sorted(result_ids)


# MAIN: iterate conexões
for i, row in df_conn.iterrows():
    title_raw = row.get("title") if "title" in df_conn.columns else row.get("Title") or row.get("titulo")
    if pd.isna(title_raw):
        erros.append({"linha Excel": i+2, "issue": "sem title", "row": i+2})
        continue
    title_key = normalize(title_raw)
    resourceID = map_pp.get(title_key)
    if resourceID is None:
        erros.append({"linha Excel": i+2, "issue": "policy title not found in public_policies.xlsx", "title": title_raw})
        continue

    species_info_col = row.get("speciesInformationColumn") if "speciesInformationColumn" in df_conn.columns else row.get("speciesInformationColumn".lower())
    biologicalLink = row.get("biologicalLink")

    # resolve species IDs
    resolved_ids = resolve_species_ids(species_info_col, biologicalLink)

    if not resolved_ids:
        # better error messaging: if species_info_col is a column name, say value not found; else say no species resolved
        nk = normalize(str(species_info_col)) if pd.notna(species_info_col) else None
        if nk and nk in cols_norm:
            erros.append({"linha Excel": i+2, "issue": "no species matched that column=value", "column": species_info_col, "value": biologicalLink, "title": title_raw})
        else:
            erros.append({"linha Excel": i+2, "issue": "no species resolved from speciesInformationColumn", "speciesInformationColumn": species_info_col, "title": title_raw})
        continue

    # create inserts
    for sid in resolved_ids:
        inserts.append(f"INSERT INTO public_policies_species (resourceID, speciesID) VALUES ({resourceID}, {sid});")

# write outputs
out_sql = os.path.join(out_dir, "inserts_public_policies_species.sql")
with open(out_sql, "w", encoding="utf-8") as f:
    f.write("\n".join(inserts))

if erros:
    pd.DataFrame(erros).to_csv(os.path.join(out_dir, "erros_pp_species.csv"), index=False, encoding="utf-8")

print("Done.")
print("Inserts written:", len(inserts))
print("Errors written:", len(erros), "->", os.path.join(out_dir, "erros_pp_species.csv"))
